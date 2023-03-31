import openpyxl
import psycopg2
from loguru import logger
import json


class DBHandler:
    def __init__(self):
        self.conn = None
        self.cur = None
        self.commands = {
            "create": """DROP TABLE IF EXISTS data; 
               CREATE TABLE data (
                    number VARCHAR(100),
                    series VARCHAR(100),
                    issue_date VARCHAR(100),
                    ending_date VARCHAR(100),
                    subject VARCHAR(100),
                    communication_type VARCHAR(100),
                    organization VARCHAR(100),
                    place VARCHAR(255),
                    latitude VARCHAR(100),
                    longitude VARCHAR(100),
                    category VARCHAR(100),
                    usage VARCHAR(100),
                    RES_name VARCHAR(100),
                    MAC VARCHAR(100),
                    call_sign VARCHAR(100),
                    network_id VARCHAR(100),
                    azimuth VARCHAR(100),
                    KA_name VARCHAR(100),
                    KA_place VARCHAR(100),
                    power VARCHAR(100),
                    rec_freq VARCHAR(255),
                    trans_freq VARCHAR(255),
                    formula VARCHAR(100),
                    class VARCHAR(100),
                    status VARCHAR(100)
                    );""",
            "show": """SELECT * FROM data""",
            "insert": """INSERT INTO data(
                          number, series, issue_date, ending_date, subject, communication_type, organization, place, 
                          latitude, longitude, category, usage, RES_name, MAC, call_sign, network_id, azimuth, KA_name, 
                          KA_place, power, rec_freq, trans_freq, formula, class, status) VALUES (
                          %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"""
        }

        with open('config.json', 'r') as config_file:
            self.config_data = json.load(config_file)

        self.connect()


    def read_xlsx(self):
        dataframe = openpyxl.load_workbook(self.config_data["filename"])
        dataframe1 = dataframe.active
        table = [[] for i in range(dataframe1.max_row)]
        for row in range(0, dataframe1.max_row):
            for col in dataframe1.iter_cols(0, dataframe1.max_column):
                if col[row].value and col[row].value != '':
                    table[row].append(col[row].value)
                else:
                    table[row].append(None)
        table.remove(table[0])
        return table

    def connect(self):
        try:
            self.conn = psycopg2.connect(database=self.config_data["database"], user=self.config_data["user"],
                                         password=self.config_data["password"], host=self.config_data["host"],
                                         port=self.config_data["port"])
            self.cur = self.conn.cursor()
        except (Exception, psycopg2.DatabaseError) as error:
            logger.info(error)

    def create_table(self):
        if not self.conn:
            self.connect()
        try:
            self.cur.execute(self.commands["create"])
            self.conn.commit()
        except (Exception, psycopg2.DatabaseError) as error:
            logger.info(error)
            self.connect()

    def show_table(self):
        if not self.conn:
            self.connect()
        try:
            self.cur.execute(self.commands["show"])
            column_names = [desc[0] for desc in self.cur.description]
            for i in column_names:
                print("|" + i + "|", end="")
            print("\n"+"--"*90)
            for i in self.cur.fetchall():
                print(i)
                print("--" * 90)

            self.conn.commit()
        except (Exception, psycopg2.DatabaseError) as error:
            logger.info(error)
            self.connect()

    def fill_table(self):
        if not self.conn:
            self.connect()
        if not self.conn:
            self.connect()
        try:
            table = self.read_xlsx()
            self.cur.executemany(self.commands["insert"], table)
            self.conn.commit()

        except (Exception, psycopg2.DatabaseError) as error:
            logger.info(error)
            self.connect()


if __name__ == "__main__":
    db = DBHandler()
    db.create_table()
    db.fill_table()
    db.show_table()

