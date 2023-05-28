"""Replace data from .xlsx file to PostgreSQL database

"""
import json
import os
import openpyxl
import psycopg2
from loguru import logger
from tools.sql_commands import COMMANDS, create, insert, drop


class DBHandler:
    """Class for operating with DB and table

    """

    def __init__(self, config, data_file=None, args=None):
        self.conn = None
        self.cur = None
        self.columns = None
        self.filename = data_file
        with open(os.path.join('./config', config), 'r', encoding='UTF-8') as config_file:
            default_data = json.load(config_file)
            if args is not None:
                args_dict = vars(args)
                for key, val in args_dict.items():
                    if not val:
                        args_dict[key] = default_data[key]
                self.config_data = args_dict
            else:
                self.config_data = default_data
        logger.info(f'Connecting with params: {self.config_data}.....')
        self.table = self.read_xlsx()
        self.connect()

    def read_xlsx(self):
        """Reads data from .xlsx file

        Returns:
            list: data from .xlsx in python's list of lists

        """
        if not self.filename:
            logger.error('No datafile')
            return None
        workbook = openpyxl.load_workbook(self.filename)
        worksheet = workbook.active
        table = [[] for i in range(worksheet.max_row)]
        for row in range(1, worksheet.max_row):
            for cell in worksheet[row]:
                if cell.value and cell.value != '':
                    table[row].append(cell.value)
                else:
                    table[row].append('None')

        for i in table:
            if not i:
                table.remove(i)
        self.columns = table[0]
        logger.info(self.columns)
        try:
            int(table[0][0])
        except ValueError:
            table.pop(0)
        return table

    def connect(self):
        """Connect to postgres database

        """
        logger.info("Connecting to DB....")
        try:
            self.conn = psycopg2.connect(database=self.config_data["database"],
                                         user=self.config_data["user"],
                                         password=self.config_data["password"],
                                         host=self.config_data["host"],
                                         port=self.config_data["port"])
            self.cur = self.conn.cursor()
            logger.info("Success!")
        except psycopg2.DatabaseError as error:
            logger.error(error)

    def drop_table(self):
        logger.info("Creating table....")
        if not self.conn:
            self.connect()
        try:
            self.cur.execute(drop(self.config_data["table"]))
            self.conn.commit()
            logger.info("Success!")
        except psycopg2.DatabaseError as error:
            logger.info(error)
            self.connect()

    def create_table(self):
        """Creating postgres table based on .xlsx table

        """
        logger.info("Creating table....")
        if not self.conn:
            self.connect()
        try:
            self.cur.execute(create(self.config_data["table"].strip('\''), self.columns))
            self.conn.commit()
            logger.info("Success!")
        except psycopg2.DatabaseError as error:
            logger.info(error)
            self.connect()

    def show_table(self):
        """Debug function - prints db content

        """
        try:
            self.cur.execute(COMMANDS["show"], self.config_data["table"])
            column_names = [desc[0] for desc in self.cur.description]
            for i in column_names:
                print("|" + i + "|", end="")
            print("\n" + "--" * 90)
            for i in self.cur.fetchall():
                print(i)
                print("--" * 90)

            self.conn.commit()
        except psycopg2.DatabaseError as error:
            logger.info(error)
            self.connect()

    def fill_table(self):
        """Put data from .xlsx to db

        """
        if not self.filename:
            logger.error('No datafile')
            exit(1)
        logger.info("Inserting data to DB....")
        if not self.conn:
            self.connect()
        try:
            self.cur.execute(insert(self.config_data["table"].strip('\''), self.table))
            self.conn.commit()
            logger.info("Success!")
        except psycopg2.DatabaseError as error:
            logger.info(error)
            self.connect()

    def get_tables(self):
        """Put data from .xlsx to db

                """
        logger.info("Inserting data to DB....")
        if not self.conn:
            self.connect()
        try:
            self.cur.execute(COMMANDS["tables"])
            self.conn.commit()
            logger.info("Success!")
            return self.cur.fetchall()
        except psycopg2.DatabaseError as error:
            logger.info(error)
            self.connect()

    def run(self):
        if not self.filename:
            logger.error('No datafile')
            exit(1)
        if self.config_data["overwrite"]:
            self.drop_table()
        self.create_table()
        self.fill_table()
